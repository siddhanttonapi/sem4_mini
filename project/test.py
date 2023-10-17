import tkinter as tk
from tkinter import ttk
import sqlite3
import pandas as pd
from tkinter import filedialog
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
from tkinter import messagebox
import os
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import csv
from fpdf import FPDF

# Connect to the database.....
conn = sqlite3.connect('test2.db')
c = conn.cursor()

# Create the login window
root = tk.Tk()
root.title('Login Page')
root.geometry('400x300')
root.configure(bg='#17202A')  # Set background color

# Define the fonts
title_font = ('Arial', 24, 'bold')
label_font = ('Arial', 14)
entry_font = ('Arial', 14)
button_font = ('Arial', 16, 'bold')


# Passing prototypes
def open_subject_window(subject):
    pass
def admindashboard(id):
    pass
def studentdashboard(id):
    pass
def open_subject_window_student(subject):
    pass
def convert_tables_to_quiz():
    pass







# Create a label with the icon-like image
icon_image = tk.PhotoImage(file='usericon.png')
icon_image = icon_image.subsample(50, 50)  # Adjust the size of the icon
icon_label = tk.Label(root, image=icon_image)
icon_label.place(relx=0.5, rely=0.1, anchor='center')  # Adjust the position of the icon

# Add the label "Quiz Management System"
quiz_label = tk.Label(root, text="Quiz Management System", font=title_font, bg='#17202A', fg='white')
quiz_label.place(relx=0.5, rely=0.20, anchor='center')

# Change the font and background color of the MIS label and entry
mis_label = tk.Label(root, text='MIS', font=label_font, bg='#17202A', fg='white')
mis_label.place(relx=0.25, rely=0.3, anchor='center')


mis_entry = tk.Entry(root, font=entry_font)
mis_entry.place(relx=0.75, rely=0.3, anchor='center')

# Change the font and background color of the password label and entry
password_label = tk.Label(root, text='Password', font=label_font, bg='#17202A', fg='white')
password_label.place(relx=0.25, rely=0.4, anchor='center')

password_entry = tk.Entry(root, show='*', font=entry_font)
password_entry.place(relx=0.75, rely=0.4, anchor='center')
password_label.config(borderwidth=0)

# Change the font, background color, and foreground color of the login button
login_button = tk.Button(root, text='Login', font=button_font, bg='#4CAF50', fg='white')
login_button.place(relx=0.5, rely=0.6, anchor='center')

# Change the font, background color, and foreground color of the register button
reg_button = tk.Button(root, text='Register', font=button_font, bg='#4CAF50', fg='white')
reg_button.place(relx=0.8, rely=0.8, anchor='center')

# Change the font, background color, and foreground color of the admin button
admin_button = tk.Button(root, text='Admin', font=button_font, bg='#4CAF50', fg='white')
admin_button.place(relx=0.2, rely=0.8, anchor='center')


def show_student_marks(id):
    # Connect to the database
    conn = sqlite3.connect('test2.db')
    c = conn.cursor()

    # Retrieve table names from the database
    c.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = c.fetchall()

    # Create a new window
    window = tk.Tk()
    window.title('Student Marks')
    window.attributes('-fullscreen', True)  # Open in full screen mode
    window.configure(bg='#17202A')  # Set background color

    # Define the button style
    button_style = {
        'bg': '#3498DB',  # Blue button background color
        'fg': '#FFFFFF',  # White button text color
        'font': ('Arial', 14),
        'width': 20,
        'height': 2,
        'relief': 'raised',
        'bd': 0  # Border width
    }

    # Create a function for displaying table marks
    def display_table_marks(table_name, connection):
        # Create a new window for displaying the table and marks
        table_window = tk.Toplevel(window)
        table_window.title(f'{table_name} - Marks')
        table_window.geometry('600x400')

        # Create a cursor for the provided database connection
        cursor = connection.cursor()

        # Fetch data from the table
        cursor.execute(f"SELECT * FROM {table_name}")
        data = cursor.fetchall()

        # Define custom column names
        column_names = ['MIS', 'Marks']

        # Create a table view using a Tkinter Treeview widget
        table_view = ttk.Treeview(table_window, show='headings', style='Custom.Treeview')

        # Configure the Treeview columns
        table_view['columns'] = tuple(range(len(column_names)))

        # Format the Treeview columns
        for i, column in enumerate(column_names):
            table_view.column(i, width=100, anchor='center')
            table_view.heading(i, text=column)

        # Insert data into the Treeview
        for row in data:
            table_view.insert('', 'end', values=row)

        # Pack the Treeview widget
        table_view.pack(fill='both', expand=True)

    search_query = f"SELECT * FROM adminpaper WHERE ID = ?"
    c.execute(search_query, (id,))
    papers = c.fetchall()

    for paper in papers:
        paper_name = paper[1] + "_marks"
        paper_button = tk.Button(window, text=paper_name, command=lambda name=paper_name: display_table_marks(name, conn), **button_style)
        paper_button.pack(pady=5)

    # Run the main event loop
    window.mainloop()



def export_data(id):
    conn = sqlite3.connect("test2.db")
    c = conn.cursor()

    # Fetch table names ending with "_marks"
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%_marks'")
    tables = c.fetchall()

    # Create a tkinter window
    root = tk.Tk()
    root.title("Export Data")
    root.attributes('-fullscreen', True)  # Open in full screen mode
    root.configure(bg='#17202A')  # Set background color

    # Define the button style
    button_style = {
        'bg': '#3498DB',  # Blue button background color
        'fg': '#FFFFFF',  # White button text color
        'font': ('Arial', 14),
        'width': 20,
        'height': 2,
        'relief': 'raised',
        'bd': 0  # Border width
    }

    def export_table(table_name, id):
        # Create a new database connection for exporting the table data
        export_conn = sqlite3.connect("test2.db")
        export_c = export_conn.cursor()

        # Fetch data from the selected table
        export_c.execute(f"SELECT * FROM {table_name}")
        data = export_c.fetchall()

        if data:
            # Create a new window for format selection
            format_window = tk.Toplevel(root)
            format_window.title("Export Format")
            format_window.configure(bg="#17202A")  # Set background color

            def export_csv():
                # Export data to a CSV file
                filename = f"{table_name}.csv"
                with open(filename, 'w', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow(["MIS", "Marks"])
                    writer.writerows(data)

                # Display export success message
                messagebox.showinfo("Export Successful", f"Data from table '{table_name}' exported successfully.")

            def export_pdf():
                # Export data to a PDF file
                filename = f"{table_name}.pdf"
                pdf = FPDF()
                pdf.add_page()

                # Define column widths and alignment
                col_width = pdf.w / 3.5
                font_size = 12
                col_height = font_size + 2

                # Set font and table header
                pdf.set_font("Arial", "B", font_size)
                pdf.cell(col_width, col_height, "MIS", border=1)
                pdf.cell(col_width, col_height, "Marks", border=1)
                pdf.ln(col_height)

                # Set font for table content
                pdf.set_font("Arial", "", font_size)

                # Add data rows
                for row in data:
                    for item in row:
                        pdf.cell(col_width, col_height, str(item), border=1)
                    pdf.ln(col_height)

                pdf.output(filename)

                # Display export success message
                messagebox.showinfo("Export Successful", f"Data from table '{table_name}' exported successfully.")

            # Define the button style
            button_style = {
                'bg': '#3498DB',  # Blue button background color
                'fg': '#FFFFFF',  # White button text color
                'font': ('Arial', 14),
                'width': 20,
                'height': 2,
                'relief': 'raised',
                'bd': 0  # Border width
            }

            # Create buttons for format selection with the defined style
            csv_button = tk.Button(format_window, text="Export as CSV", command=export_csv, **button_style)
            csv_button.pack(padx=20, pady=10)

            pdf_button = tk.Button(format_window, text="Export as PDF", command=export_pdf, **button_style)
            pdf_button.pack(padx=20, pady=10)

        else:
            messagebox.showinfo("No Data", f"No data found in table '{table_name}'.")

        export_conn.close()


    search_query = f"SELECT * FROM adminpaper WHERE ID = ?"
    c.execute(search_query, (id,))
    papers = c.fetchall()

    for paper in papers:
        paper_name = paper[1] + "_marks"
        button = tk.Button(root, text=paper_name, command=lambda name=paper_name: export_table(name,id), **button_style)
        button.pack(pady=5)


    # Run the tkinter event loop for the window
    root.mainloop()

    conn.close()





def adminlogin():
    try:
        root.destroy()
        admin_window = tk.Tk()
        admin_window.title('Admin Login')
        admin_window.geometry('400x300')
        admin_window.configure(bg='#17202A')  # Dark background

        # Widgets with color adjustments
        titleadmin_label = tk.Label(admin_window, text='Login', font=('Arial', 40), fg='#E74C3C', bg='#17202A')  # Red title
        id_label = tk.Label(admin_window, text='ID', font=('Arial', 18), fg='white', bg='#17202A')
        id_entry = tk.Entry(admin_window, font=('Arial', 18), bg='#2C3E50', fg='white')  # Dark background for entry
        pw_label = tk.Label(admin_window, text='Password', font=('Arial', 18), fg='white', bg='#17202A')
        pw_entry = tk.Entry(admin_window, show='*', font=('Arial', 18), bg='#2C3E50', fg='white')
        loginadmin_button = tk.Button(admin_window, text='Login', font=('Arial', 20), bg='#3498db', fg='white')  # Blue button
        regadmin_button = tk.Button(admin_window, text='Register', font=('Arial', 20), bg='#3498db', fg='white')  # Blue button

        # Layout
        titleadmin_label.place(relx=0.5, rely=0.1, anchor='center')
        id_label.place(relx=0.3, rely=0.3, anchor='center')
        id_entry.place(relx=0.7, rely=0.3, anchor='center')
        pw_label.place(relx=0.3, rely=0.4, anchor='center')
        pw_entry.place(relx=0.7, rely=0.4, anchor='center')
        loginadmin_button.place(relx=0.5, rely=0.6, anchor='center')
        regadmin_button.place(relx=0.8, rely=0.8, anchor='center')

        def loginadmin():
            try:
                admin_id = id_entry.get()
                admin_pw = pw_entry.get()

                # Database check
                c.execute("SELECT * FROM ADMINLOGIN WHERE ID=? AND PW=?", (admin_id, admin_pw))
                row = c.fetchone()

                if row is not None:
                    admin_window.destroy()
                    admindashboard2(admin_id)
                else:
                    messagebox.showerror("Login Failed", "Incorrect ID or password. Please try again.")
                    pw_entry.delete(0, tk.END)
            except Exception as e:
                messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

        # Button configurations
        loginadmin_button.configure(command=loginadmin)
        regadmin_button.configure(command=lambda: print("Register Button Clicked"))

        admin_window.eval('tk::PlaceWindow . center')
        admin_window.attributes("-fullscreen", True)
        admin_window.attributes('-zoomed', True)
        admin_window.mainloop()

    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")




def registerpageadmin():
    # Create the register window
    regadmin_window = tk.Tk()
    regadmin_window.title('Register')
    regadmin_window.geometry('400x300')
    regadmin_window.attributes('-fullscreen', True)  # Open in full screen mode
    regadmin_window.configure(bg='#17202A')  # Set background color

    # Define the fonts
    title_font = ('Arial', 40, 'bold')  # Increase font size
    label_font = ('Arial', 18)
    entry_font = ('Arial', 18)
    button_font = ('Arial', 20, 'bold')  # Increase font size

    # Create the widgets
    title2_label = tk.Label(regadmin_window, text='Register', font=title_font, fg='#E74C3C', bg='#17202A')
    idreg_label = tk.Label(regadmin_window, text='Admin ID', font=label_font, fg='#ECF0F1', bg='#17202A')
    idreg_entry = tk.Entry(regadmin_window, font=entry_font)
    pwreg_label = tk.Label(regadmin_window, text='Password', font=label_font, fg='#ECF0F1', bg='#17202A')
    pwreg_entry = tk.Entry(regadmin_window, show='*', font=entry_font)
    registeradmin_button = tk.Button(regadmin_window, text='Register', font=button_font, bg='#4CAF50', fg='#ECF0F1')

    # Define the layout
    title2_label.place(relx=0.5, rely=0.1, anchor='center')
    idreg_label.place(relx=0.3, rely=0.3, anchor='center')
    idreg_entry.place(relx=0.7, rely=0.3, anchor='center')
    pwreg_label.place(relx=0.3, rely=0.4, anchor='center')
    pwreg_entry.place(relx=0.7, rely=0.4, anchor='center')
    registeradmin_button.place(relx=0.5, rely=0.6, anchor='center')

    def registeradmin():
        idreg = idreg_entry.get()
        pwreg = pwreg_entry.get()
        # Insert the new student into the database
        c.execute("INSERT INTO ADMINLOGIN (ID,PW) VALUES (?, ?)", (idreg, pwreg))
        conn.commit()
        # Show a success message and clear the entry widgets
        success_label = tk.Label(regadmin_window, text='SUCCESS', fg='green', font=label_font, bg='#17202A')
        success_label.place(relx=0.5, rely=0.8, anchor='center')
        regadmin_window.after(2000, regadmin_window.destroy)  # destroys the register_window after 2 seconds

    registeradmin_button.configure(command=registeradmin)


    regadmin_window.mainloop()





# Define the login function
def login():
    # Get the MIS and password from the entry widgets
    mis= mis_entry.get()
    password = password_entry.get()

    # Check the credentials against the database
    c.execute("SELECT * FROM STUDENTLOGIN WHERE MIS=? AND Password=?", (mis, password))
    row = c.fetchone()
    if row is not None:
        studentdashboard(mis)
    else:
        # If the login fails, show an error message and prompt the user to try again
        error_label = tk.Label(root, text='Incorrect MIS or password. Please try again.', fg='red', font=label_font, bg='#f2f2f2')
        error_label.place(relx=0.5, rely=0.7, anchor='center')
        password_entry.delete(0, tk.END)

def registerpage():
    # Create the register window
    register_window = tk.Tk()
    register_window.title('Register')
    register_window.geometry('400x300')
    register_window.attributes('-fullscreen', True)  # Open in full screen mode
    register_window.configure(bg='#17202A')  # Set background color

    # Define the fonts
    title_font = ('Arial', 40, 'bold')  # Increase font size
    label_font = ('Arial', 18)
    entry_font = ('Arial', 18)
    button_font = ('Arial', 20, 'bold')  # Increase font size


    # Create the widgets
    title2_label = tk.Label(register_window, text='Register', font=title_font, fg='#E74C3C', bg='#17202A')
    mis2_label = tk.Label(register_window, text='MIS', font=label_font, fg='white',bg='#17202A')
    mis2_entry = tk.Entry(register_window, font=entry_font)
    name2_label = tk.Label(register_window, text='Name', font=label_font, bg='#17202A', fg='white')
    name2_entry = tk.Entry(register_window, font=entry_font)
    password2_label = tk.Label(register_window, text='Password', font=label_font, bg='#17202A', fg='white')
    password2_entry = tk.Entry(register_window, show='*', font=entry_font)
    register2_button = tk.Button(register_window, text='Register', font=button_font, bg='#4CAF50', fg='white')

    # Define the layout
    title2_label.place(relx=0.5, rely=0.1, anchor='center')
    mis2_label.place(relx=0.3, rely=0.3, anchor='center')
    mis2_entry.place(relx=0.7, rely=0.3, anchor='center')
    name2_label.place(relx=0.3, rely=0.4, anchor='center')
    name2_entry.place(relx=0.7, rely=0.4, anchor='center')
    password2_label.place(relx=0.3, rely=0.5, anchor='center')
    password2_entry.place(relx=0.7, rely=0.5, anchor='center')
    register2_button.place(relx=0.5, rely=0.6, anchor='center')


    def register():
        mis = mis2_entry.get()
        name = name2_entry.get()
        password = password2_entry.get()

        # Validate the MIS format
        if not mis[0].isalpha():
            error_label = tk.Label(register_window, text='MIS must start with a letter', fg='red', font=label_font, bg='#17202A')
            error_label.place(relx=0.5, rely=0.8, anchor='center')
            return

        # Insert the new student into the database
        c.execute("INSERT INTO STUDENTLOGIN (MIS, Name, Password) VALUES (?, ?, ?)", (mis, name, password))
        conn.commit()

        # Create a new table for the student
        c.execute(f'CREATE TABLE "{mis}" (Quizname TEXT NOT NULL, Quizscore REAL NOT NULL)')
        conn.commit()

        # Show a success message and clear the entry widgets
        success_label = tk.Label(register_window, text='SUCCESS', fg='green', font=label_font, bg='#f2f2f2')
        success_label.place(relx=0.5, rely=0.8, anchor='center')
        register_window.after(2000, register_window.destroy)  # destroys the register_window after 2 seconds

    register2_button.configure(command=register)
    register_window.attributes("-fullscreen", True)
    register_window.attributes('-zoomed', True)


# Bind the login function to the login button
login_button.configure(command=login)
reg_button.configure(command=registerpage)
admin_button.configure(command=adminlogin)

def open_subject_window(subject,id):
    # Create a new window
    subject_window = tk.Toplevel()
    subject_window.title(subject)
    subject_window.geometry('400x300')
    subject_window.configure(bg='#17202A')  # Dark blue background color

    # Define the font
    title_font = ('Arial', 24, 'bold')

    # Create a label for the title
    title_label = tk.Label(subject_window, text=subject, font=title_font, bg='#17202A', fg='#FFFFFF')  # White text color
    title_label.pack(pady=50)

    # Define the button style
    button_style = {
        'bg': '#3498DB',  # Blue button background color
        'fg': '#FFFFFF',  # White button text color
        'font': ('Arial', 14),
        'width': 20,
        'height': 2,
        'relief': 'raised',
        'bd': 0  # Border width
    }

    # Define the button hover style
    button_hover_style = {
        'bg': '#2980B9',  # Darker blue background color on hover
        'fg': '#FFFFFF',  # White button text color on hover
    }

    def insert_data(file_path, id):
        # Read the data from the file
        data = pd.read_excel(file_path, header=None)

        # Get the name of the table from the file name
        workbook = file_path.split('/')[-1].split('.')[0]
        marks_table = f"{workbook}_marks"

        # Connect to the database
        conn = sqlite3.connect('test2.db')
        c = conn.cursor()

        # Check if the marks table already exists
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (marks_table,))
        result = c.fetchone()
        if result:
            messagebox.showerror("Error", "Quiz with same name already exists")
            conn.close()
            return

        # Drop the existing table if it exists
        drop_table_query = f"DROP TABLE IF EXISTS {workbook}"
        c.execute(drop_table_query)

        # Create the table in the database
        create_table_query = f"CREATE TABLE {workbook} (question TEXT, option1 TEXT, option2 TEXT, option3 TEXT, option4 TEXT, answer TEXT)"
        c.execute(create_table_query)

        # Loop through the rows in the Excel file and insert them into the table
        for row in data.values.tolist():
            insert_query = f"INSERT INTO {workbook} (question, option1, option2, option3, option4, answer) VALUES (?, ?, ?, ?, ?, ?)"
            c.execute(insert_query, row)

        # Create a table called {workbook}_marks with two columns: ID and Papername
        create_marks_table_query = f"CREATE TABLE {marks_table} (MIS TEXT NOT NULL, marks TEXT NOT NULL)"
        c.execute(create_marks_table_query)

        # Insert the ID and name into the adminpaper table
        insert_marks_query = f"INSERT INTO adminpaper (ID, Papername) VALUES (?, ?)"
        c.execute(insert_marks_query, (id, workbook))

        # Commit the changes and close the connection
        conn.commit()
        conn.close()

        # Display a success message to the user
        messagebox.showinfo('Success', 'Data inserted into the database')


    # Define the function to handle file selection
    def select_file():
        file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
        if file_path:
            insert_data(file_path,id)

    # Create the button to select a file
    select_file_button = tk.Button(subject_window, text='Select File', command=select_file, **button_style)
    select_file_button.pack(pady=10)

    # Apply hover effect to the button
    select_file_button.bind('<Enter>', lambda event: select_file_button.config(**button_hover_style))
    select_file_button.bind('<Leave>', lambda event: select_file_button.config(**button_style))



def admindashboard2(id):

    # Create a new window
    dashboard_window = tk.Tk()
    dashboard_window.title('Dashboard')
    dashboard_window.geometry('800x600')
    dashboard_window.configure(bg='#17202A')
    dashboard_window.attributes("-fullscreen", True)
    dashboard_window.attributes('-zoomed', True)

    # Define the fonts
    title_font = ('Arial', 24, 'bold')
    welcome_font = ('Arial', 18, 'bold')
    button_font = ('Arial', 16, 'bold')

    # Create the header
    header_frame = tk.Frame(dashboard_window, bg='#C0C0C0')
    header_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
    header_frame.columnconfigure(0, weight=1)

    # Add the welcome label
    welcome_label = tk.Label(header_frame, text="Welcome, " + id, font=welcome_font, bg='#C0C0C0', fg='black', padx=500, pady=15)
    welcome_label.grid(row=0, column=0, sticky='w')

    # Add a spacer
    spacer_label = tk.Label(header_frame, text="", bg='#C0C0C0')
    spacer_label.grid(row=0, column=1, sticky='e')

    # Create the widgets
    subject1_button = tk.Button(dashboard_window, text='VCPDE', font=button_font, bg='#3498DB', fg='white', command=lambda: open_subject_window('VCPDE',id))
    subject2_button = tk.Button(dashboard_window, text='DSA-II', font=button_font, bg='#3498DB', fg='white', command=lambda: open_subject_window('DSA',id))
    subject3_button = tk.Button(dashboard_window, text='MPT', font=button_font, bg='#3498DB', fg='white', command=lambda: open_subject_window('MPT',id))
    subject4_button = tk.Button(dashboard_window, text='DC', font=button_font, bg='#3498DB', fg='white', command=lambda: open_subject_window('DC',id))
    
    # Create the graph buttons
    graph1_button = tk.Button(dashboard_window, text='Export Marks', font=button_font, bg='#3498DB', fg='white',command=lambda: export_data(id))
    graph2_button = tk.Button(dashboard_window, text='View Student Analytics', font=button_font, bg='#3498DB', fg='white',command=lambda: show_student_marks(id))
    
    # Add the widgets to the window
    subject1_button.grid(row=1, column=0, padx=20, pady=20, sticky='nsew')
    subject2_button.grid(row=1, column=1, padx=20, pady=20, sticky='nsew')
    subject3_button.grid(row=2, column=0, padx=20, pady=20, sticky='nsew')
    subject4_button.grid(row=2, column=1, padx=20, pady=20, sticky='nsew')
    graph1_button.grid(row=3, column=0, padx=20, pady=20, sticky='nsew')
    graph2_button.grid(row=3, column=1, padx=20, pady=20, sticky='nsew')

    # Set the grid layout parameters
    for i in range(4):
        dashboard_window.rowconfigure(i, weight=1)
    for i in range(2):
        dashboard_window.columnconfigure(i, weight=1)

    # Center the welcome label
    welcome_label.grid_configure(padx=(200, 0))

    # Add styles to the buttons
    dashboard_window.option_add('*Button.activeBackground', '#3e8e41')
    dashboard_window.option_add('*Button.activeForeground', 'white')
    dashboard_window.option_add('*Button.highlightBackground', '#7be583')
    dashboard_window.option_add('*Button.highlightColor', 'white')
    dashboard_window.option_add('*Button.relief', 'raised')
    dashboard_window.option_add('*Button.borderWidth', '2')

    # Add styles to the header
    header_frame.configure(highlightbackground='#C0C0C0', highlightcolor='#C0C0C0', highlightthickness=1)

    # Add styles to the welcome label
    welcome_label.configure(anchor='center')

    # Add styles to the subject buttons
    subject1_button.configure(width=20, height=10)
    subject2_button.configure(width=20, height=10)
    subject3_button.configure(width=20, height=10)
    subject4_button.configure(width=20, height=10)

    # Add styles to the graph buttons
    graph1_button.configure(width=20, height=5)
    graph2_button.configure(width=20, height=5)

    # Add a style for the spacer label
    spacer_label.configure(width=30)

    # Add a footer
    footer_frame = tk.Frame(dashboard_window, bg='#C0C0C0')
    footer_frame.grid(row=4, column=0, columnspan=2, sticky='nsew')
    footer_label = tk.Label(footer_frame, text='© 2023 All Rights Reserved. | Created by Siddhant & Pushkar', font=('Arial', 10, 'italic'), bg='#C0C0C0', fg='black')
    footer_label.pack(pady=10)

    # Start the window
    dashboard_window.mainloop()

def studentdashboard(id):

    def create_quiz_graph(id):
        # Connect to the database and extract data
        conn = sqlite3.connect('test2.db')
        c = conn.cursor()
        c.execute(f"SELECT Quizname, Quizscore FROM {id}")
        rows = c.fetchall()
        conn.close()

        # Sort the rows by Quizname in alphabetical order
        rows.sort(key=lambda x: x[0])

        # Convert the Quizscore from string to float
        scores = [float(row[1]) for row in rows]

        # Create a bar graph
        fig = plt.figure(figsize=(8, 5))
        ax = fig.add_subplot(111)

        # Set styles for the bar graph
        ax.bar([row[0] for row in rows], scores, color='#3498DB', edgecolor='black')  # Blue bars with black edges

        # Set y limit to 100
        ax.set_ylim([0, 100])

        # Set styles for the axes labels and tick labels
        ax.set_xlabel('Quiz Name', fontweight='bold', fontsize=12, labelpad=10)
        ax.set_ylabel('Score', fontweight='bold', fontsize=12, labelpad=10)
        ax.tick_params(axis='x', labelrotation=45, labelsize=10)
        ax.tick_params(axis='y', labelsize=10)

        # Set title styles
        ax.set_title('Quiz Scores', fontweight='bold', fontsize=16, pad=20)

        # Create a Tkinter window and add the graph to it
        root = tk.Tk()
        root.title("Bar Graph")
        root.attributes('-fullscreen', True)  # Open in fullscreen

        # Configure the Tkinter window to occupy the whole window
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)

        canvas = FigureCanvasTkAgg(fig, master=root)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)  # Fill and expand the canvas to occupy the whole window

        root.mainloop()


    def export_table_as_pdf(table_name):
        # Create a new database connection for exporting the table data
        export_conn = sqlite3.connect("test2.db")
        export_c = export_conn.cursor()

        # Fetch data from the selected table
        export_c.execute(f"SELECT * FROM {table_name}")
        data = export_c.fetchall()

        if data:
            # Export data to a PDF file
            filename = f"{table_name}.pdf"
            pdf = FPDF()
            pdf.add_page()

            # Define column widths and alignment
            col_width = pdf.w / 3.5
            font_size = 12
            col_height = font_size + 2

            # Set font and table header
            pdf.set_font("Arial", "B", font_size)
            pdf.cell(col_width, col_height, "MIS", border=1)
            pdf.cell(col_width, col_height, "Marks", border=1)
            pdf.ln(col_height)

            # Set font for table content
            pdf.set_font("Arial", "", font_size)

            # Add data rows
            for row in data:
                for item in row:
                    pdf.cell(col_width, col_height, str(item), border=1)
                pdf.ln(col_height)

            pdf.output(filename)

            # Create a new window for the download message
            download_window = tk.Toplevel()
            download_window.title("Download Successful")

            # Create a label with the download message
            download_label = tk.Label(download_window, text="Marks downloaded.", font=('Arial', 14))
            download_label.pack(padx=20, pady=10)

        else:
            messagebox.showinfo("No Data", f"No data found in table '{table_name}'.")

        export_conn.close()

    def export_table_as_csv(table_name):
        # Create a new database connection for exporting the table data
        export_conn = sqlite3.connect("test2.db")
        export_c = export_conn.cursor()

        # Fetch data from the selected table
        export_c.execute(f"SELECT * FROM {table_name}")
        data = export_c.fetchall()

        if data:
            # Export data to a CSV file
            filename = f"{table_name}.csv"
            with open(filename, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(["MIS", "Marks"])
                writer.writerows(data)

            # Create a new window for the download message
            download_window = tk.Toplevel()
            download_window.title("Download Successful")

            # Create a label with the download message
            download_label = tk.Label(download_window, text="Marks downloaded as CSV.", font=('Arial', 14))
            download_label.pack(padx=20, pady=10)

        else:
            messagebox.showinfo("No Data", f"No data found in table '{table_name}'.")

        export_conn.close()


    # Create a new window
    dashboard_window = tk.Tk()
    dashboard_window.title('Dashboard')
    dashboard_window.geometry('800x600')
    dashboard_window.configure(bg='#17202A')  # Dark blue background
    dashboard_window.attributes("-fullscreen", True)
    dashboard_window.attributes('-zoomed', True)

    # Define the fonts
    title_font = ('Arial', 24, 'bold')
    welcome_font = ('Arial', 18, 'bold')
    button_font = ('Arial', 16, 'bold')

    # Create the header
    header_frame = tk.Frame(dashboard_window, bg='silver')  # Silver background
    header_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
    header_frame.columnconfigure(0, weight=1)

    # Add the welcome label
    welcome_label = tk.Label(header_frame, text="Welcome, " + id, font=welcome_font, bg='silver', fg='black',padx=500)  # Black text
    welcome_label.grid(row=0, column=0, sticky='w')

    # Add a spacer
    spacer_label = tk.Label(header_frame, text="", bg='silver')
    spacer_label.grid(row=0, column=1, sticky='e')

    # Create the widgets
    subject1_button = tk.Button(dashboard_window, text='VCPDE', font=button_font, bg='#3498DB', fg='white',command=lambda: convert_tables_to_quiz(id, "vcpde"))  # Blue button
    subject2_button = tk.Button(dashboard_window, text='DSA-II', font=button_font, bg='#3498DB', fg='white',command=lambda: convert_tables_to_quiz(id, "dsa"))  # Blue button
    subject3_button = tk.Button(dashboard_window, text='MPT', font=button_font, bg='#3498DB', fg='white',command=lambda: convert_tables_to_quiz(id, "mpt"))  # Blue button
    subject4_button = tk.Button(dashboard_window, text='DC', font=button_font, bg='#3498DB', fg='white',command=lambda: convert_tables_to_quiz(id, "dc"))  # Blue button

    # Create the graph buttons
    graph1_button = tk.Button(dashboard_window, text='View graph', font=button_font, bg='#3498DB', fg='white', command=lambda: create_quiz_graph(id))  # Blue button
    graph2_button = tk.Button(dashboard_window, text='Download Marks as PDF', font=button_font, bg='#3498DB', fg='white',command=lambda: export_table_as_pdf(id))  # Blue button

    # Add the widgets to the window
    subject1_button.grid(row=1, column=0, padx=20, pady=20, sticky='nsew')
    subject2_button.grid(row=1, column=1, padx=20, pady=20, sticky='nsew')
    subject3_button.grid(row=2, column=0, padx=20, pady=20, sticky='nsew')
    subject4_button.grid(row=2, column=1, padx=20, pady=20, sticky='nsew')
    graph1_button.grid(row=3, column=0, padx=20, pady=20, sticky='nsew')
    graph2_button.grid(row=3, column=1, padx=20, pady=20, sticky='nsew')

    # Set the grid layout parameters
    for i in range(4):
        dashboard_window.rowconfigure(i, weight=1)
    for i in range(2):
        dashboard_window.columnconfigure(i, weight=1)

    # Center the welcome label
    welcome_label.grid_configure(padx=(200, 0))

    # Add styles to the buttons
    dashboard_window.option_add('*Button.activeBackground', '#0000FF')  # Blue active background
    dashboard_window.option_add('*Button.activeForeground', 'white')
    dashboard_window.option_add('*Button.highlightBackground', '#0000FF')  # Blue highlight background
    dashboard_window.option_add('*Button.highlightColor', 'white')
    dashboard_window.option_add('*Button.relief', 'raised')
    dashboard_window.option_add('*Button.borderWidth', '2')

    # Add styles to the header
    header_frame.configure(highlightbackground='silver', highlightcolor='silver', highlightthickness=1)

    # Add styles to the welcome label
    welcome_label.configure(anchor='center')

    # Add styles to the subject buttons
    subject1_button.configure(width=20, height=10)
    subject2_button.configure(width=20, height=10)
    subject3_button.configure(width=20, height=10)
    subject4_button.configure(width=20, height=10)

    # Add styles to the graph buttons
    graph1_button.configure(width=20, height=5)
    graph2_button.configure(width=20, height=5)

    # Add a footer
    footer_frame = tk.Frame(dashboard_window, bg='silver')  # Silver background
    footer_frame.grid(row=4, column=0, columnspan=2, sticky='nsew')
    footer_label = tk.Label(footer_frame, text='© 2023 All Rights Reserved. | Created by Siddhant & Pushkar',
                            font=('Arial', 10, 'italic'), bg='silver', fg='black')  # Black text
    footer_label.pack(pady=10)

    # Start the window
    dashboard_window.mainloop()




def convert_tables_to_quiz(id, subject_name):
    subject_name2 = subject_name.upper()  # Capitalize the subject_name argument
    conn = sqlite3.connect("test2.db")
    c = conn.cursor()
    c.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '{subject_name}%'")
    tables = c.fetchall()
    quiz_list = []
    for table in tables:
        c.execute(f"PRAGMA table_info({table[0]})")
        columns = c.fetchall()
        if len(columns) == 6:
            c.execute(f"SELECT * FROM {table[0]}")
            rows = c.fetchall()
            quiz = []
            for row in rows:
                question = row[0]
                options = [row[1], row[2], row[3], row[4]]
                answer = row[5]
                quiz.append((question, options, answer))
            quiz_list.append((table[0], quiz))  # Append the table name along with the quiz
    conn.close()

    # Create a tkinter window
    root = tk.Tk()
    root.title("Quizzes")

    # Set window properties
    root.attributes('-fullscreen', True)  # Open in fullscreen
    root.configure(bg="#17202A")  # Set dark blue background color

    # Custom styles
    quiz_button_style = {
        'bg': '#3498DB',  # Blue button background color
        'fg': '#FFFFFF',  # White button text color
        'font': ('Arial', 14, 'bold'),
        'width': 20,
        'height': 2,
        'relief': 'raised',
        'bd': 0  # Border width
    }

    # Create a header label
    header_label = tk.Label(root, text=f"{subject_name2} Quizzes", font=('Arial', 20, 'bold'), bg="#17202A", fg="#FFFFFF")
    header_label.pack(pady=20)

    for i, (quiz_name, quiz) in enumerate(quiz_list):
        button = tk.Button(
            root,
            text=quiz_name,
            command=lambda quiz=quiz, quiz_name=quiz_name: start_quiz(quiz, id, quiz_name),
            **quiz_button_style
        )
        button.pack(pady=10)

    root.mainloop()


def start_quiz(quiz, id, quiz_name):
    # Define a function to grade the quiz

    # Check if the quiz has already been given
    conn = sqlite3.connect("test2.db")
    c = conn.cursor()
    table_name = str(id)  # Create table name from id argument
    c.execute(f"SELECT Quizname FROM {table_name}")
    entries = c.fetchall()
    conn.close()

    # Check if the quiz_name exists in the entries
    if any(quiz_name in entry for entry in entries):
        # Create a new window for the message
        message_window = tk.Toplevel(root)
        message_window.title("Quiz Already Given")

        # Create a label with the message
        message_label = tk.Label(message_window, text="You have already given this quiz!", font=('Arial', 14))
        message_label.pack(padx=20, pady=10)

        return
    
    def grade_quiz():
        # Get the user's answers
        user_answers = []
        for question in quiz:
            answer = answer_vars[question[0]].get()
            user_answers.append(answer)

        # Grade the quiz
        num_correct = 0
        for i, question in enumerate(quiz):
            if user_answers[i] == question[2]:  # Compare the selected option with the answer
                num_correct += 1
        score = 100 * num_correct / len(quiz)

        # Insert the quiz results into the database
        conn = sqlite3.connect("test2.db")
        c = conn.cursor()
        table_name = str(id)  # Create table name from id argument
        c.execute(f"INSERT INTO {table_name} (Quizname, Quizscore) VALUES (?, ?)", (quiz_name, score))
        conn.commit()
        conn.close()

        # Update the quiz results in the database
        conn = sqlite3.connect("test2.db")
        c = conn.cursor()
        table_name = f"{quiz_name}_marks"  # Create table name with quiz name followed by "_marks"
        c.execute(f"INSERT INTO {table_name} (MIS, marks) VALUES (?, ?)", (id, score))
        # Update the existing record in the table

        conn.commit()
        conn.close()
        quiz_window.destroy()
        # Create a new window for the quiz results
        results_window = tk.Toplevel(root)
        results_window.title("Quiz Results")

        # Create a label with the quiz results
        results_label = tk.Label(results_window, text=f"You scored {score:.2f} % ({num_correct}/{len(quiz)})!", font=('Arial', 14))
        results_label.pack(padx=20, pady=10)

    # Create a new window for the quiz
    quiz_window = tk.Toplevel(root)
    quiz_window.title("Quiz")

    # Make the quiz window fullscreen
    quiz_window.attributes("-fullscreen", True)

    # Get the screen width and height
    screen_width = quiz_window.winfo_screenwidth()
    screen_height = quiz_window.winfo_screenheight()

    # Calculate the center position of the quiz frame
    quiz_frame_width = 400  # Adjust as needed
    quiz_frame_height = 300  # Adjust as needed
    x = (screen_width - quiz_frame_width) // 2
    y = (screen_height - quiz_frame_height) // 2

    # Create a canvas to contain the quiz questions and options
    canvas = tk.Canvas(quiz_window)
    canvas.pack(side="left", fill="both", expand=True)

    # Add a scrollbar to the canvas
    scrollbar = tk.Scrollbar(quiz_window, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Create a frame for the quiz questions and options on the canvas
    quiz_frame = tk.Frame(canvas)
    canvas.create_window((x, y), window=quiz_frame, anchor="nw")  # Set the position of the quiz frame

    # Create a dictionary to store the answer variables
    answer_vars = {}

    # Define font styles
    question_font = ("Arial", 16, "bold")
    option_font = ("Arial", 14)

    # Display each quiz question and options
    for i, question in enumerate(quiz):
        # Create a label for the question
        question_label = tk.Label(quiz_frame, text=f"{i+1}. {question[0]}", anchor="center", font=question_font)
        question_label.pack(pady=10)

        # Create variables for the answer options
        answer_vars[question[0]] = tk.StringVar(value="")  # Initialize as empty string

        # Create radio buttons for the answer options
        for j, option in enumerate(question[1]):
            option_button = tk.Radiobutton(quiz_frame, text=option, variable=answer_vars[question[0]], value=option, font=option_font)
            option_button.pack(anchor="w")

    # Create a submit button
    submit_button = tk.Button(quiz_frame, text="Submit", command=grade_quiz, font=option_font)
    submit_button.pack(pady=20)

    # Update the canvas to show all quiz questions and options
    quiz_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    # Configure the canvas scrolling
    canvas.configure(scrollregion=canvas.bbox("all"))
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=quiz_frame, anchor="nw")

    # Create a label for the timer
    time_left = 900  # 15 minutes in seconds
    timer_label = tk.Label(quiz_window, text=f"Time left: {time_left // 60:02d}:{time_left % 60:02d}", font=option_font)
    timer_label.pack()

    def update_timer():
        nonlocal time_left  # Use the time_left variable defined in the parent function

        # Decrement time_left and update the timer label
        if time_left > 0:
            time_left -= 1
            timer_label.config(text=f"Time left: {time_left // 60:02d}:{time_left % 60:02d}")
        else:
            # If time runs out, grade the quiz and close the window
            grade_quiz()
            quiz_window.destroy()

        # Schedule the next call to update_timer after 1 second
        quiz_window.after(1000, update_timer)

    update_timer()

# Set the minimum size of the window
root.minsize(600, 800)
root.attributes('-fullscreen', True)  # Open in full screen mode
# Center the window on the screen
root.eval('tk::PlaceWindow . center')
# Run the main loop
root.mainloop()

# Close the database connection when the program ends
conn.close() 


# Connect to the database.....
# Connect to the database.....
# Connect to the database.....
# Connect to the database.....
# Connect to the database.....